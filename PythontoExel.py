import openpyxl as xl
from openpyxl.styles import Font, numbers

wb = xl.Workbook()

ws = wb.active

ws.title = 'First Sheet'

wb.create_sheet(index = 1, title= 'Second Sheet')





# write content to a cell

ws['A1'] = 'Inovice'

fontobj = Font(name = 'Times New Roman', size = 24, bold =  True)

ws['A1'].font = fontobj

ws['A2'] =  'Tires'
ws['A3'] =  'Brakes'
ws['A4'] =  'Alignent'

ws['B2'] = 450
ws['B3'] = 225.50
ws['B4'] = 150

ws['A8'] = 'Total'
ws['A8'].font = fontobj

ws.merge_cells('A1:B1')

ws.column_dimensions['A'].width =  25
ws['B8'] = '=SUM(B2:B7)'

# write produce report to second sheet in current workbook

write_sheet = wb['Second Sheet']

read_wb = xl.load_workbook('ProduceReport.xlsx')
read_ws = read_wb['ProduceReport']

fontobj = Font(name = 'Times New Roman', size = 18, bold =  True)

counter = 1
for currentrow in read_ws.iter_rows(min_row=1):
    write_sheet['A' + str(counter)] = currentrow[0].value
    write_sheet['B' + str(counter)] = currentrow[1].value
    write_sheet['C' + str(counter)] = currentrow[2].value
    write_sheet['D' + str(counter)] = currentrow[3].value
    counter += 1

counter += 1 

write_sheet['B'+ str(counter)] = "Total"
write_sheet['B'+ str(counter)].font = fontobj

write_sheet['C'+ str(counter)] = f"=SUM(C2:C{counter - 2})"

write_sheet['D'+ str(counter)] = f"=SUM(D2:D{counter - 2})"


counter +=1


write_sheet['B'+ str(counter)] = "Average"
write_sheet['B'+ str(counter)].font =  fontobj

write_sheet['C'+ str(counter)] = f"=AVERAGE(C2:C{counter - 3})"


write_sheet['D'+ str(counter)] = f"=AVERAGE(D2:D{counter - 3})"



write_sheet.column_dimensions['A'].width =  16
write_sheet.column_dimensions['B'].width =  16
write_sheet.column_dimensions['C'].width =  16
write_sheet.column_dimensions['D'].width =  16

for cell in write_sheet['C:C']:
    cell.number_format =  '"$ "#,##0.00'

for cell in write_sheet['D:D']:
    cell.number_format =  '"$ "#,##0.00'


wb.save('PythontoExcel.xlsx')